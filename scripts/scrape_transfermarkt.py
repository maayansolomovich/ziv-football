#!/usr/bin/env python3
"""
Scrape player data from Transfermarkt and update the players Excel file.

Usage:
    python3 scripts/scrape_transfermarkt.py [path/to/players.xlsx]

Requirements:
    pip install playwright openpyxl
    python -m playwright install --channel chrome   # or: playwright install chromium

The script reads the Excel file, finds rows that have a Transfermarkt URL
but are missing scraped data (DOB, position, foot, team), fetches each page,
and writes the results back to the same file.

Excel expected columns (Hebrew headers):
    שם בעברית, שם באנגלית, עמדה, תאריך לידה, גיל, גובה (ס"מ), רגל חזקה,
    קבוצה נוכחית, חתום (כן/לא), תחילת חוזה, סיום חוזה, קישור Transfermarkt
"""

import re
import sys
import time
from pathlib import Path

import openpyxl
from playwright.sync_api import sync_playwright

# ---------------------------------------------------------------------------
# Translation maps
# ---------------------------------------------------------------------------

POSITION_MAP = {
    "Attacking Midfield": "קשר תוקפני",
    "Central Midfield": "קשר מרכזי",
    "Defensive Midfield": "קשר הגנתי",
    "Right Midfield": "קשר ימין",
    "Left Midfield": "קשר שמאל",
    "Goalkeeper": "שוער",
    "Left Winger": "כנף שמאל",
    "Right Winger": "כנף ימין",
    "Left-Back": "מגן שמאל",
    "Right-Back": "מגן ימין",
    "Centre-Back": "מגן מרכזי",
    "Centre-Forward": "חלוץ מרכזי",
    "Second Striker": "תוקף שני",
    "Sweeper": "ליברו",
}

TEAM_MAP = {
    "Hapoel Haifa": "הפועל חיפה",
    "Hapoel Tel Aviv": "הפועל תל אביב",
    "Hapoel Beer Sheva": "הפועל באר שבע",
    "Hapoel Raanana": "הפועל רעננה",
    "Hapoel Jerusalem": "הפועל ירושלים",
    "Hapoel Kfar Shalem": "הפועל כפר שלם",
    "Hapoel Petah Tikva": "הפועל פתח תקווה",
    "Hapoel Afula": "הפועל עפולה",
    "Maccabi Tel Aviv": "מכבי תל אביב",
    "Maccabi Haifa": "מכבי חיפה",
    "Maccabi Netanya": "מכבי נתניה",
    "Maccabi Herzliya": "מכבי הרצליה",
    "Maccabi Jaffa": "מכבי יפו",
    "Maccabi Kiryat Malachi": "מכבי קריית מלאכי",
    "Maccabi Ahi Nazareth": "מכבי אחי נצרת",
    "Maccabi Bnei Reineh": "מכבי בני ריינה",
    "Maccabi Petah Tikva": "מכבי פתח תקווה",
    "Beitar Jerusalem": "בית\"ר ירושלים",
    "Bnei Yehuda": "בני יהודה",
    "Ironi Kiryat Shmona": "עירוני קריית שמונה",
    "Ironi Tiberias": "עירוני טבריה",
    "Ironi Nesher": "עירוני נשר",
    "FC Ashdod": "מ.ס. אשדוד",
    "Sektzia Nes Tziona": "סקציה נס ציונה",
    "Ahi Nazareth": "אחי נצרת",
    "Without Club": "ללא קבוצה",
}

FOOT_MAP = {"right": "ימין", "left": "שמאל", "both": "שתיים"}


# ---------------------------------------------------------------------------
# Scraping
# ---------------------------------------------------------------------------

def scrape_player(page, url: str) -> dict:
    page.goto(url, timeout=30_000)
    time.sleep(1.5)

    info = page.evaluate("""
        () => {
            const labels = document.querySelectorAll(
                '.info-table__content--regular'
            );
            const pairs = {};
            labels.forEach(el => {
                const label = el.textContent.trim().replace(':', '');
                const next = el.nextElementSibling;
                if (next) pairs[label] = next.textContent.trim();
            });
            return pairs;
        }
    """)

    dob_raw = info.get("Date of birth/Age", "")
    dob_m = re.search(r"(\d{2}/\d{2}/\d{4})", dob_raw)
    dob = dob_m.group(1) if dob_m else ""
    height_raw = info.get("Height", "")
    height_m = re.search(r"(\d)[,\.](\d+)", height_raw)
    height = int(height_m.group(1) + height_m.group(2).ljust(2, "0")[:2]) if height_m else ""

    pos_raw = info.get("Position", "")
    pos_en = pos_raw.split(" - ")[-1] if " - " in pos_raw else pos_raw
    pos = POSITION_MAP.get(pos_en, pos_en)

    foot_en = info.get("Foot", "").lower().strip()
    foot = FOOT_MAP.get(foot_en, "ימין")

    team_en = info.get("Current club", "")
    team = TEAM_MAP.get(team_en, team_en)

    return {
        "dob": dob,
        "age": age,
        "height": height,
        "pos":            pos,
        "foot":           foot,
        "team":           team,
        "contract_start": info.get("Joined", ""),
        "contract_end":   info.get("Contract expires", ""),
    }


# ---------------------------------------------------------------------------
# Excel update
# ---------------------------------------------------------------------------

def col_index(headers: dict, name: str) -> int | None:
    return headers.get(name)


def update_excel(path: Path) -> None:
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # Build header → column-index map
    headers = {ws.cell(row=1, column=i).value: i for i in range(1, ws.max_column + 1)}

    C = {
        "pos":            col_index(headers, "עמדה"),
        "dob":            col_index(headers, "תאריך לידה"),
        "height":         col_index(headers, 'גובה (ס"מ)'),
        "foot":           col_index(headers, "רגל חזקה"),
        "team":           col_index(headers, "קבוצה נוכחית"),
        "signed":         col_index(headers, "חתום (כן/לא)"),
        "contract_start": col_index(headers, "תחילת חוזה"),
        "contract_end":   col_index(headers, "סיום חוזה"),
        "tm_url":         col_index(headers, "קישור Transfermarkt"),
    }

    # Collect rows that have a TM URL
    rows_to_scrape = []
    for ri in range(2, ws.max_row + 1):
        url_cell = ws.cell(row=ri, column=C["tm_url"]) if C["tm_url"] else None
        url = url_cell.value if url_cell else None
        if url and url.startswith("https://www.transfermarkt.com"):
            rows_to_scrape.append((ri, url))

    if not rows_to_scrape:
        print("No Transfermarkt URLs found — nothing to scrape.")
        return

    print(f"Found {len(rows_to_scrape)} players with TM URLs. Scraping...")

    with sync_playwright() as p:
        browser = p.chromium.launch(
            channel="chrome",
            headless=True,
            args=["--no-sandbox", "--disable-dev-shm-usage"],
        )
        page = browser.new_page(
            user_agent=(
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/125.0.0.0 Safari/537.36"
            )
        )

        for ri, url in rows_to_scrape:
            name = ws.cell(row=ri, column=headers.get("שם בעברית", 1)).value or ""
            print(f"  [{ri-1}/{len(rows_to_scrape)}] {name} ... ", end="", flush=True)
            try:
                data = scrape_player(page, url)

                def set_if_empty(col_key, value):
                    if C.get(col_key) and value:
                        cell = ws.cell(row=ri, column=C[col_key])
                        # Always overwrite — re-running updates stale data
                        cell.value = value

                set_if_empty("dob", data["dob"])
                set_if_empty("height", data["height"])
                set_if_empty("pos", data["pos"])
                set_if_empty("foot", data["foot"])
                set_if_empty("team", data["team"])
                set_if_empty("contract_start", data["contract_start"])
                set_if_empty("contract_end", data["contract_end"])

                if C.get("signed"):
                    ws.cell(row=ri, column=C["signed"]).value = (
                        "כן" if data["team"] and data["team"] != "ללא קבוצה" else "לא"
                    )

                print(f"DOB={data['dob']}  גובה={data['height']}  רגל={data['foot']}  קבוצה={data['team']}")
            except Exception as e:
                print(f"ERROR — {e}")

        browser.close()

    wb.save(path)
    print(f"\nSaved: {path}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    default = Path.home() / "Downloads" / "players_profile_full.xlsx"
    target = Path(sys.argv[1]) if len(sys.argv) > 1 else default

    if not target.exists():
        print(f"File not found: {target}")
        sys.exit(1)

    update_excel(target)
